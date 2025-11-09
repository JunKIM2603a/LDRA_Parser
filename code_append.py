from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QPushButton, QMessageBox
from PySide6.QtCore import QDir, QStringListModel, QTimer
from PySide6.QtUiTools import QUiLoader
from mainwindow import Ui_MainWindow

import os
import re
from collections import defaultdict, Counter
import copy
import shutil
import openpyxl

class MainWindow(QMainWindow, Ui_MainWindow):
    # 초기화
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)  
        self.folder_path_src = None
        self.folder_path_htm = None
        self.Stage = None
        self.Rule = None
        self.pattern = None

        self.setupUi(self)
        self.initUI()
        # 정규표현식
        # self.pattern = r"/\/\*MISRA_\d{1,3}[A-Z]_\d{1,3}\*\/"

    # UI 생성시 작업
    def initUI(self):
        self.setWindowTitle("Code Append - LDRA")
        self.label_Status.setText("대기 중")

        # QListView에 사용할 모델 설정
        self.model_src_files = QStringListModel()
        self.listView_src_files.setModel(self.model_src_files)
        self.btn_src_folder.clicked.connect(self.open_src)

        self.model_result_files = QStringListModel()
        self.listView_result_files.setModel(self.model_result_files)
        self.btn_htm_folder.clicked.connect(self.open_htm)

        self.btn_start.clicked.connect(self.code)
        self.model_result_rules = QStringListModel()
        self.listView_result_rules.setModel(self.model_result_rules)


        self.btn_Restore.clicked.connect(self.delete_annotation)

        # "Extract Detection Lines" 버튼 연결
        self.btn_extract_data.clicked.connect(self.extract_detection_lines)

        # QTreeWidget 생성
        self.treeWidget_result.setColumnCount(1)  # 열 개수 (파일명 / 에러번호)
        self.treeWidget_result.setHeaderLabels(["파일별 에러번호"])  # 헤더 설정
        


    ## ----------------------------------------------------------------------------------------------
    # 기능: 소스파일 열기 - 1
    def open_src(self):
        self.folder_path_src = QFileDialog.getExistingDirectory(self, "폴더 선택", "")
        if self.folder_path_src:
            self.label_src_folder_path.setText(f"선택한 폴더: {self.folder_path_src}")
            self.Show_src_files()
        else:
            self.label_src_folder_path.setText(f"폴더를 선택하지 않았습니다.")

    # 기능: 소스파일 열기 - 2
    def Show_src_files(self):
        if self.folder_path_src:
            # QDir 객체 생성
            dir = QDir(self.folder_path_src)
            
            # 디렉토리 내에서 .cpp와 .hpp 파일만 필터링
            dir.setNameFilters(["*.cpp", "*.hpp"])
        
            # .cpp, .hpp 파일 목록 가져오기
            files = dir.entryList(QDir.Files)  

            sorted_list = sorted(files)
            sorted_list.insert(0, f"파일 개수: {len(sorted_list)}")
            print(sorted_list)
            # 모델에 파일 목록 설정
            self.model_src_files.setStringList(sorted_list)

    # 기능: htm파일 열기 - 1
    def open_htm(self):
        self.folder_path_htm = QFileDialog.getExistingDirectory(self, "폴더 선택", "")
        if self.folder_path_htm:
            self.label_htm_folder_path.setText(f"선택한 폴더: {self.folder_path_htm}")
            print(self.folder_path_htm)
            # self.code()
            vio = self.get_misra_violations(self.folder_path_htm)
            self.Show_result_files(vio)
        else:
            self.label_htm_folder_path.setText(f"폴더를 선택하지 않았습니다.")

    # 기능: htm파일 열기 - 2
    def Show_result_files(self, vio):
        if self.folder_path_htm:
            # 모델에 파일 목록 설정

            # 데이터에서 파일명만 추출
            files = set(entry[3] for entry in vio)  # 중복 제거를 위해 set 사용
            sorted_list = sorted(files)
            sorted_list.insert(0,f"파일 개수: {len(sorted_list)}")
            print(sorted_list)
            self.model_result_files.setStringList(sorted_list)

    def extract_detection_lines(self):
        if not self.folder_path_htm:
            QMessageBox.warning(self, "경고", "먼저 'LDRA htm 폴더 열기'를 실행하세요.")
            return

        folder_name = os.path.basename(self.folder_path_htm)
        excel_file_name = f"{folder_name}_extract.xlsx"
        excel_file_path = os.path.join(os.path.dirname(self.folder_path_htm), excel_file_name)

        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # 기본 시트 제거

            prefix = folder_name.split('_tbwrkfls')[0]

            for file_name in os.listdir(self.folder_path_htm):
                if not file_name.endswith('.htm'):
                    continue

                rule_match = re.search(rf"{prefix}_link_popup(\d+[A-Za-z])\.htm", file_name)
                if not rule_match:
                    continue

                rule_name = rule_match.group(1)
                sheet_name = rule_name.replace(" ", "_")
                if len(sheet_name) > 31:  # 시트 이름 길이 제한
                    sheet_name = sheet_name[:31]
                
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(title=sheet_name)

                file_path = os.path.join(self.folder_path_htm, file_name)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                # 1. 룰 설명 및 위반 사항 추출
                # desc_match = re.search(r'<h3>(.*?)</h3>', content)                                         
                # description_html = desc_match.group(1).strip() if desc_match else "Description not found"
                # description_text = re.sub(r'<.*?>', '', description_html).strip()  
                all_h3_matches = re.findall(r'<h3>(.*?)</h3>', content)
                description_html = ""
                for h3_content in all_h3_matches:
                    h3_text = re.sub(r'<.*?>', '', h3_content).strip()
                    if re.match(r'^\d+\s*-\s*', h3_text):
                        description_html = h3_content
                        break
                
                if not description_html and all_h3_matches:
                    description_html = all_h3_matches[0] # Fallback

                description_text_with_prefix = re.sub(r'<.*?>', '', description_html).strip()
                description_text = re.sub(r'^\d+\s*-\s*', '', description_text_with_prefix).strip()
                
                spaced_rule_name = re.sub(r'(\d+)([A-Za-z]+)', r'\1 \2', rule_name)

                final_violations = []
                if rule_name == '38S':
                    # Special handling for 38S
                    s38_locations = re.findall(r'<a href="ldra://editor/\?File=[^&]*\\([^&]+)&line=(\d+)">', content)
                    for file, line in s38_locations:
                        final_violations.append((file, line))
                else:
                    # General case for other rules
                    violations = re.findall(r"<b>Location</b>\s*:\s*<a href\s*=\s*'.*?'>([^<]+)</a>(?: - <a href=\"ldra://editor/.*?&Line=(\d+)\">\d+</a>)?", content)
                    
                    if rule_name in ['45D', '128D', '70D', '91D', '3X']:
                        if violations:
                            final_violations.append(violations[0])
                    elif rule_name in ['49D', '68X']:
                        if violations:
                            final_violations.append(violations[-1])
                    else:
                        final_violations = violations
                
                detection_count = len(final_violations)

                # 2. 헤더 작성
                sheet.cell(row=1, column=1, value=spaced_rule_name)
                sheet.cell(row=1, column=2, value=description_text)
                sheet.cell(row=1, column=3, value=detection_count)

                # 3. 시트에 데이터 쓰기
                row_num = 2
                for func, line in final_violations:
                    sheet.cell(row=row_num, column=1, value=func.strip())
                    if line:
                        sheet.cell(row=row_num, column=2, value=line.strip())
                    else:
                        sheet.cell(row=row_num, column=2, value="Reference to function location")
                    row_num += 1

            if not workbook.sheetnames:
                workbook.create_sheet(title="No Data")

            workbook.save(excel_file_path)
            QMessageBox.information(self, "성공", f"'{excel_file_path}' 파일이 생성되었습니다.")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 생성 중 오류가 발생했습니다:\n{e}")

    ## ----------------------------------------------------------------------------------------------
    # 기능: .htm 파일에서 MISRA 룰 위반 정보를 추출
    def get_misra_violations(self, htm_folder):
        # 경로에서 _tbwrkfls 앞의 문자열 추출
        folder_name = os.path.basename(htm_folder)  # 경로에서 폴더 이름만 추출
        prefix = folder_name.split('_tbwrkfls')[0]  # '_tbwrkfls' 앞의 문자열을 추출: KDDX_SDI_MISRA_tbwrkfls 의 경우 prefix = KDDX_SDI_MISRA

        print(f"Set: {prefix}")

        """
        주어진 폴더 내의 모든 .htm 파일을 읽고 MISRA 룰 위반 라인 정보를 추출합니다.
        :param htm_folder: HTML 파일들이 위치한 폴더 경로
        :return: 룰 번호와 해당하는 라인 번호의 리스트
        """
        violations = []
        Cnt_Total = 0
        Cnt_Rule = 0
        for file_name in os.listdir(htm_folder):
            if file_name.endswith('.htm'):
                file_path = os.path.join(htm_folder, file_name)
                with open(file_path, 'r', encoding='utf-8') as file:    # 룰 별로 검출된 내용이 있는 파일 열기
                    # htm_folder 내부의 개별 검출내용 htm 파일 확인 ex> KDDX_SDI_MISRA_link_popup1D.htm
                    search_param = rf"{prefix}_link_popup(\d+[A-Za-z])\.htm"
                    rule_match = re.search(search_param, file_name) # 파일에서 검출내용 검색
                    if rule_match: # if htm_folder 내부의 개별 검출내용 htm 파일 별 내용이 있다면.
                        # group(0) 전체 매칭된 문자열을 반환
                        # group(1) 정규표현식에서 첫 번째 매칭된 값을 반환
                        # group(2) 정규표현식에서 두 번째 매칭된 값을 반환
                        # KDDX_SDI_MISRA_link_popup(\d+[A-Za-z])\.htm
                        # KDDX_SDI_MISRA_link_popup1D.htm -> rule_name_number = 1D
                        rule_name_number = rule_match.group(1)
                        Cnt_Rule += 1
                        # 코드에서 MISRA 룰 위반 라인 정보 찾기
                        Cnt_Rule_Total = 0
                        file_lines = 0
                        while True:
                            # 한 줄씩 읽기
                            content = file.readline()
                            file_lines += 1
                            if not content: # 다 읽으면 다음파일 읽기
                                print(file.name, file_lines)
                                break
                            line_matches = None
                            match = None

                            # 검출에러 1개당, 위치 여러개 검출되면, 앞에 검출된 위치로 사용
                            if rule_name_number == '45D' or rule_name_number == '128D' or rule_name_number == '70D' or rule_name_number == '91D' or rule_name_number == '3X':
                                # <a href="ldra://editor/?File=C:\Users\user\Desktop\vx_wksp\KDDX_SDI\SDI_Vpx3C3_CWE\SDI_Vpx3C3\CEthernetTask.cpp&Line=4022">4022</a>,
                                line_matches = None
                                line_matches = re.findall(r'<a href="ldra://editor/\?File=([^\&]+)&Line=(\d+)">(\d+)</a>', content)
                                match = line_matches[0] if line_matches else None   

                            # 검출에러 1개당, 위치 여러개 검출되면, 뒤에 검출된 위치로 사용
                            elif rule_name_number == '49D' or rule_name_number == '68X':
                                line_matches = None
                                line_matches = re.findall(r'<a href="ldra://editor/\?File=([^\&]+)&Line=(\d+)">(\d+)</a>', content)
                                if len(line_matches) > 1:   
                                    match = line_matches[1] if line_matches else None   
                                else:
                                    match = line_matches[0] if line_matches else None

                            # 검출에러 1개당, 1개 검출되면, 검출된 위치로 사용
                            else:
                                line_matches = None
                                line_matches = re.findall(r'<a href="ldra://editor/\?File=([^\&]+)&Line=(\d+)">(\d+)</a><br>', content)
                                match = line_matches[0] if line_matches else None  

                            if match:
                                # match[0]: 소스코드 path
                                # match[1]: 소스코드 별 검출 라인
                                # cpp_file: 파일 명
                                cpp_file = os.path.basename(match[0])
                                # cpp_file = match[0]
                                line_number = int(match[1])
                                # 0: 전체 검출 개수 카운트, 
                                # 1: 검출 Rule 파일 카운트, 
                                # 2: Rule 별 검출 개수 카운트, 
                                # 3: 파일명
                                # 4: line_number, 
                                # 5: rule_name_number
                                Cnt_Rule_Total += 1
                                Cnt_Total += 1
                                violations.append((Cnt_Total, Cnt_Rule, Cnt_Rule_Total, cpp_file, line_number, rule_name_number))   
        return violations

    ## ----------------------------------------------------------------------------------------------
    def add_misra_comments_to_cpp(self, cpp_folder, violations):
        """
        C++ 파일을 읽고 MISRA 룰 위반 라인에 주석을 추가합니다.
        :param cpp_folder: C++ 파일들이 위치한 폴더 경로
        :param violations: MISRA 룰 위반 라인 정보 (룰 번호와 해당 라인 번호)
        """
        num_violations = len(violations)
        violations_list = []
        cnt_file = 0
        cnt_violations = 0
        # for file_name in os.listdir(cpp_folder): 
        for root, dirs, file_names in os.walk(cpp_folder): 
            # if file_name.endswith('.cpp'):
            # print(root, dirs, file_names, type(file_names))
            for file_name in file_names:
                # print(file_name, root, dirs)
                # print(os.path.join(root, file_name))

                if file_name.endswith('.cpp') or file_name.endswith('.hpp') or file_name.endswith('.h'):
                    cnt_file += 1
                    # print(cnt_file, file_name)
                    # file_path = os.path.join(cpp_folder, file_name)
                    file_path = os.path.join(root, file_name)


                    # 파일 이름과 확장자 분리
                    filename, ext = os.path.splitext(file_name)
                    # 백업 파일 이름 생성
                    backup_file_name = f"{filename}_bak{ext}"
                    # 복사하여 백업 파일 만들기
                    backup_file_path = os.path.join(root, backup_file_name)
                    shutil.copyfile(file_path, backup_file_path)
                    print(f"백업 파일이 생성되었습니다: {backup_file_path}")

                    with open(backup_file_path, 'w', encoding='utf-8') as backup_file:
                        line_num = 0
                        with open(file_path, 'r+', encoding='utf-8') as file:
                            while True:
                                line_one = file.readline()
                                line_num += 1
                                if not line_one: # 다 읽으면 다음파일 읽기
                                    print(file.name, line_one)
                                    break
                                # backup_file.seek(0)
                                for index, item in enumerate(violations):
                                    # print(len(violations), len(violations) - index, index, file_name, item[3], item[4])
                                    # 0: 전체 검출 개수 카운트, 
                                    # 1: 검출 Rule 파일 카운트, 
                                    # 2: Rule 별 검출 개수 카운트, 
                                    # 3: 파일명
                                    # 4: line_number, 
                                    # 5: rule_name_number
                                    if item[3] == file_name and item[4] == line_num:
                                        # print(f"인덱스: {index}")
                                        # print(item, violations[index])
                                        comment_line = f'/*{self.Stage}_{self.Rule}_{item[5]}*/\n'
                                        # line_one = comment_line + line_one
                                        line_one = line_one + comment_line
                                        # print(line)
                                        # print(lines[i])
                                        # 파일의 처음으로 돌아가서 수정된 내용을 다시 씁니다
                                        if item[5] == '70D':
                                            print(item[5], item[3], item[4])
                                        # violations.pop(index)
                                        violations_list.append(index)
                                        cnt_violations += 1

                                backup_file.write(line_one)
                                backup_file.truncate()  # 파일 크기를 현재 위치로 잘라냄 (기존에 남은 내용이 있으면 삭제)
                                line_one = ""
                                


                                # line_all = file.readlines()
                                # backup_file.seek(0)
                                # backup_file.writelines(line_all)  # 수정된 모든 라인을 다시 파일에 씁니다
                                # backup_file.truncate()  # 파일 크기를 현재 위치로 잘라냄 (기존에 남은 내용이 있으면 삭제)
                                # for line_num, line_one in enumerate(line_all):
                                #     # print(i, line)
                                #     # 룰 위반 라인에 주석 추가
                                #     # 인덱스를 찾는 코드
                                #     for index, item in enumerate(violations):
                                #         # print(len(violations), len(violations) - index, index, file_name, item[3], item[4])
                                #         # 0: 전체 검출 개수 카운트, 
                                #         # 1: 검출 Rule 파일 카운트, 
                                #         # 2: Rule 별 검출 개수 카운트, 
                                #         # 3: 파일명
                                #         # 4: line_number, 
                                #         # 5: rule_name_number
                                #         if item[3] == file_name and item[4] == line_num:
                                #             # print(f"인덱스: {index}")
                                #             # print(item, violations[index])
                                #             comment_line = f'/*{self.Stage}_{self.Rule}_{item[5]}*/\n'
                                #             line_one = comment_line + line_one
                                #             line_all[line_num] = line_one
                                #             # print(line)
                                #             # print(lines[i])
                                #             # 파일의 처음으로 돌아가서 수정된 내용을 다시 씁니다
                                #             if item[5] == '45D':
                                #                 print(item[5], item[3], item[4])
                                #             backup_file.seek(0)
                                #             backup_file.writelines(line_all)  # 수정된 모든 라인을 다시 파일에 씁니다
                                #             backup_file.truncate()  # 파일 크기를 현재 위치로 잘라냄 (기존에 남은 내용이 있으면 삭제)
                                #             violations.pop(index)
                                #             cnt_violations += 1

                    # 백업 파일을 원본으로 덮어쓰기
                    shutil.copyfile(backup_file_path, file_path)

                    # 백업 파일 삭제
                    os.remove(backup_file_path)

                    print(f"{backup_file_path} 내용을 {file_path}로 복원하고, {backup_file_path}은 삭제했습니다.")

                    print(cnt_file, file_name, len(violations), cnt_violations) 

        violations_list.sort(reverse=True)
        for idx in violations_list:
            if 0 <= idx < len(violations):
                violations.pop(idx)
        print("남은 데이터:", violations)
                    
        # print(cnt_violations, "/", num_violations, "(", num_violations - cnt_violations, "):", type(violations), len(violations), type(violations[0]), violations[-1])




    ## ----------------------------------------------------------------------------------------------
    # Function 2: Rule-wise detection count with sorting by S -> D -> X order and numbers ascending
    def count_rule_detections_sorted(self, violations):
        rule_count = defaultdict(int)

        # Count detections for each rule
        for _, _, rule, _, _, rule_name in violations:
            rule_count[rule_name] += 1

        # Custom order for suffixes (S -> D -> X)
        suffix_order = {'S': 0, 'D': 1, 'X': 2}

        # Sort the rules first by the suffix (S -> D -> X), then by the numeric part
        def sort_key(rule_name):
            suffix = rule_name[-1]  # Last character is the suffix
            number = int(rule_name[:-1])  # All characters except the last one are the number
            return (suffix_order[suffix], number)

        # Sort the rule names based on the defined order
        sorted_rules = sorted(rule_count.items(), key=lambda x: sort_key(x[0]))

        return sorted_rules

    # Function 3: File-wise detection count along with line_number and rule_name_number
    def count_file_detections(self, violations):
        file_count = defaultdict(lambda: defaultdict(lambda: {'count': 0, 'details': []}))
        
        for _, _, rule, file, line, rule_name in violations:
            file_count[file][rule_name]['count'] += 1
            file_count[file][rule_name]['details'].append((line, rule_name))
        
        return file_count
    
    # 기능: 검출내용 주석화
    def code(self):
        QTimer.singleShot(0, lambda: self.label_Status.setText("실행 중"))
        self.Stage = self.comboBox_Stage.currentText()    # 'DT' or 'OT'
        self.Rule = self.comboBox_Rule.currentText()    # 'MISRA' or 'CWE'
        # 1. .htm 파일에서 MISRA 룰 위반 정보를 추출
        violations = self.get_misra_violations(self.folder_path_htm)
        # 0: 전체 검출 개수 카운트, 
        # 1: 검출 Rule 파일 카운트, 
        # 2: Rule 별 검출 개수 카운트, 
        # 3: 파일명
        # 4: line_number, 
        # 5: rule_name_number
        # print("{rule_sel} Violations Found:", violations)
        # print(type(violations), len(violations), type(violations[0]), violations[-1])


        # 2. Rule-wise detection count with sorting
        violations_for_tree = copy.deepcopy(violations)
        sorted_rule_detections = self.count_rule_detections_sorted(violations)
        print("Rule-wise Detection Count (sorted by S -> D -> X and number ascending):")
        for rule, count in sorted_rule_detections:
            print(f"{rule}: {count}")
        print(f"Total: {len(violations)}")

        # 3. File-wise detection count with line number and rule_name_number
        file_detections = self.count_file_detections(violations)
        print("\nFile-wise Detection Count:")
        for file, rules in file_detections.items():
            print(f"File: {file}")
            for rule_name, data in rules.items():
                print(f"  Rule: {rule_name}, Count: {data['count']}")
                for detail in data['details']:
                    print(f"    Line: {detail[0]}, Rule: {detail[1]}")


        # 4. C++ 파일에서 룰 위반 라인에 주석을 추가
        self.add_misra_comments_to_cpp(self.folder_path_src, violations)
        # while True:
        #     self.add_misra_comments_to_cpp(self.folder_path_src, violations)
        #     if not violations:
        #         break
        print(violations)
        print(type(violations), len(violations))

        self.get_count_error_per_rule(sorted_rule_detections)
        self.get_tree_result(violations_for_tree)

        QTimer.singleShot(0, lambda: self.label_Status.setText("완료"))
        QTimer.singleShot(0, lambda: self.label_Result.setText("주석완료"))
        print("{self.Stage} {self.Rule} comments added to C++ files.")

    
    ## ----------------------------------------------------------------------------------------------
    # 기능: 주석제거
    # def delete_annotation(self):
    #     QTimer.singleShot(0, lambda: self.label_Status.setText("실행 중"))
    #     self.Stage = self.comboBox_Stage.currentText()    # 'DT' or 'OT'
    #     self.Rule = self.comboBox_Rule.currentText()    # 'MISRA' or 'CWE'
    #     # 폴더 경로가 지정되었으면
    #     if self.folder_path_src:
    #         # .cpp, .hpp 파일만 찾기
    #         files_to_process = [
    #             f for f in os.listdir(self.folder_path_src)
    #             if f.endswith(('.cpp', '.hpp'))
    #         ]
            
    #         # 각 파일을 순회하며 처리
    #         for file_name in files_to_process:
    #             file_path = os.path.join(self.folder_path_src, file_name)

    #             # 파일 읽기
    #             with open(file_path, 'r', encoding='utf-8') as file:
    #                 file_content = file.read()

    #             # 정규표현식으로 주석 제거
    #             self.pattern = rf"/\*{self.Stage}_{self.Rule}_\d{{1,3}}[A-Z]\*\/"
    #             self.pattern = self.pattern.replace("\\\\","\\")    # '\\' → '\'
    #             modified_content = re.sub(self.pattern, '', file_content)

    #             # 수정된 내용이 원본과 다르면 파일을 덮어쓰기
    #             if modified_content != file_content:
    #                 with open(file_path, 'w', encoding='utf-8') as file:
    #                     file.write(modified_content)

    #             print(f"Processed: {file_name}")
    #     QTimer.singleShot(0, lambda: self.label_Status.setText("완료"))
    #     QTimer.singleShot(0, lambda: self.label_Result.setText("주석제거"))

    def delete_annotation(self):
        QTimer.singleShot(0, lambda: self.label_Status.setText("실행 중"))
        self.Stage = self.comboBox_Stage.currentText()    # 'DT' or 'OT'
        self.Rule = self.comboBox_Rule.currentText()    # 'MISRA' or 'CWE'
        # 폴더 경로가 지정되었으면
        if self.folder_path_src:
            # .cpp, .hpp 파일만 찾기
            files_to_process = [
                f for f in os.listdir(self.folder_path_src)
                if f.endswith(('.cpp', '.hpp'))
            ]
            
            # 각 파일을 순회하며 처리
            for file_name in files_to_process:
                file_path = os.path.join(self.folder_path_src, file_name)
                print(f"Processing: {file_name}")


                with open(file_path, 'r+', encoding='utf-8') as file:
                    lines = file.readlines()

                # 정규표현식 패턴
                pattern = rf"/\*{self.Stage}_{self.Rule}_\d{{1,3}}[A-Z]\*/"
                pattern = pattern.replace("\\\\", "\\")

                modified_lines = []
                i = 0
                while i < len(lines):
                    current_line = lines[i]
                    stripped_line = current_line.rstrip('\n')

                    # 줄 전체가 패턴과 일치하면 → 삭제
                    if re.fullmatch(pattern, stripped_line):
                        # 다음 줄이 존재할 경우 들여쓰기 복원
                        if i + 1 < len(lines):
                            next_line = lines[i + 1]
                            indent = re.match(r"^(\s*)", current_line).group(1)  # 삭제된 줄의 들여쓰기
                            # trimmed_next_line = next_line.lstrip()  # 다음 줄의 선행 공백 제거
                            # lines[i + 1] = indent + trimmed_next_line  # 들여쓰기 복원
                            lines[i + 1] = indent + next_line  # 들여쓰기 복원
                        i += 1  # 현재 줄 스킵 (삭제)
                        continue

                    # 중간에 패턴이 있다면 제거
                    cleaned_line = re.sub(pattern, '', current_line)
                    modified_lines.append(cleaned_line)
                    i += 1

                # 파일 덮어쓰기
                with open(file_path, 'w', encoding='utf-8') as file:
                    file.writelines(modified_lines)

        print(f"Finished Restore")
        QTimer.singleShot(0, lambda: self.label_Status.setText("완료"))
        QTimer.singleShot(0, lambda: self.label_Result.setText("주석제거"))



    ## ----------------------------------------------------------------------------------------------
    # 기능: Rule 별 검출개수
    def get_count_error_per_rule(self, sorted_rule_detections):
        error_per_rule = []
        total_error = 0
        for rule, count in sorted_rule_detections:
            # print(f"{rule}: {count}")
            total_error += count
            error_per_rule.append(f"{rule}: {count}") 
        # print(f"Total: {len(sorted_rule_detections)}")
        error_per_rule.insert(0, f"검출 개수: {total_error}")
        self.model_result_rules.setStringList(error_per_rule)

    ## ----------------------------------------------------------------------------------------------
    # 기능: 파일의 검출목록
    def get_tree_result(self, violations):
        # 파일별 에러번호 데이터
        file_errors = self.get_file_errors(violations)
        # 트리 채우기
        self.populate_tree(file_errors)


    def get_file_errors(self, violations):
        """ 파일별 중복 없는 에러번호를 딕셔너리로 반환 """
        file_errors = defaultdict(dict)
        sorted_rule_detections = self.count_rule_detections_sorted(violations)
        rule_count_map = dict(sorted_rule_detections)  # 빠른 lookup용

        for entry in violations:
            file_name = entry[3]  # 파일 이름
            error_code = entry[5]  # 에러번호
            if error_code not in file_errors[file_name]:
                file_errors[file_name][error_code] = rule_count_map.get(error_code, 1)

        return {
            file: sorted(file_errors[file].items())
            for file in file_errors
        }

    def populate_tree(self, file_errors):
        for file_name, error_list in file_errors.items():
            file_item = QTreeWidgetItem(self.treeWidget_result)
            file_item.setText(0, f"📂 {file_name}")

            total_error = 0
            for _, count in error_list:
                total_error += count
            error_item = QTreeWidgetItem(file_item)
            error_item.setText(0, f"Total: {total_error}")   

            for error_code, count in error_list:
                error_item = QTreeWidgetItem(file_item)
                total_error += count
                error_item.setText(0, f"{error_code}: {count}")


    # def get_file_errors(self, violations):
    #     """ 파일별 중복 없는 에러번호를 딕셔너리로 반환 """
    #     file_errors = defaultdict(set)  # 중복 방지를 위해 set 사용
    #     file_error_counts = defaultdict(set) # 해당 에러의 카운트 (정수) 저장

    #     sorted_rule_detections = self.count_rule_detections_sorted(violations)
    #     rule_count_map = dict(sorted_rule_detections)  # 빠른 lookup용
        
    #     for entry in violations:
    #         file_name = entry[3]  # 파일 이름
    #         error_code = entry[5]  # 에러번호
    #         for rule, count in sorted_rule_detections:
    #             if(error_code == rule):
    #                 file_errors[file_name].add(error_code)  # 중복 없이 추가
    #                 # file_errors[file_name].add(count)  # 중복 없이 추가
    #                 file_error_counts[file_name].add(rule_count_map[error_code])

    #     # return {file: sorted(list(errors)) for file, errors in file_errors.items()}  # 정렬된 리스트 변환
    #         # 문자열과 숫자를 섞지 않고 각각 정렬해서 반환
    #     return {
    #         file: {
    #             'error_codes': sorted(file_errors[file]),
    #             'counts': sorted(file_error_counts[file])
    #         }
    #         for file in file_errors
    #     }

    # def populate_tree(self, file_errors):
    #     """ QTreeWidget을 파일별 에러번호 구조로 채움 """
    #     for file_name, error_list in file_errors.items():
    #         error_codes = error_list.get('error_codes', [])
    #         counts = error_list.get('counts', [])
    #         file_item = QTreeWidgetItem(self.treeWidget_result)  # 파일을 루트 노드로 추가
    #         file_item.setText(0, f"📂 {file_name}")  # 파일명 설정

    #         # 에러코드와 카운트가 대응되는 순서라고 가정
    #         for i, error_code in enumerate(error_codes):
    #             count = counts[i] if i < len(counts) else '?'
    #             error_item = QTreeWidgetItem(file_item)
    #             error_item.setText(0, f"{error_code}: {count}")

    #         # for error_code, count in error_list:
    #         #     error_item = QTreeWidgetItem(file_item)  # 에러번호를 자식 노드로 추가
    #         #     error_item.setText(0, f"{error_code}: {count}")



def main(): 
    app = QApplication()
    window = MainWindow()
    window.show()
    app.exec_()



if __name__ == '__main__':
    main()


                    
def a():
    # htm_folder = r'C:\LDRA_Workarea\KDDX_SDI_MISRA_tbwrkfls'  # htm 파일들이 있는 폴더 경로
    # cpp_folder = r'C:\Users\user\Desktop\vx_wksp\KDDX_SDI\SDI_Vpx3C3'  # cpp 파일들이 있는 폴더 경로
    # TEST 용도
    htm_folder = r'C:\LDRA_Workarea\SDI_Vpx3C3_20250313_1328_tbwrkfls'  # htm 파일들이 있는 폴더 경로
    cpp_folder = r'C:\Users\user\Desktop\vx_wksp\KDDX_SDI\SDI_Vpx3C3_20250313_1328\SDI_Vpx3C3'  # cpp 파일들이 있는 폴더 경로


